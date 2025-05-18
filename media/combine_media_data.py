"""
% ccm_modify_date: 2025-05-18 16:57:22 %
% ccm_author: mpegg %
% ccm_version: 43 %
% ccm_repo: https://github.com/mpegg007/TermiteTowers.git %
% ccm_branch: main %
% ccm_object_id: media/combine_media_data.py:43 %
% ccm_commit_id: 85077287515bd36c372cecb566bd8b590687d30d %
% ccm_commit_count: 43 %
% ccm_last_commit_message: move config read %
% ccm_last_commit_author: Matthew Pegg %
% ccm_last_commit_date: 2025-03-22 17:57:56 -0400 %
% ccm_file_last_modified: 2025-05-18 16:52:46 %
% ccm_file_name: combine_media_data.py %
% ccm_file_type: text/plain %
% ccm_file_encoding: us-ascii %
% ccm_file_eol: CRLF %
"""

import os
import logging
from openpyxl import Workbook
import sqlite3

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Function to fetch data from the database
def fetch_data(query, db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(query)
    data = cursor.fetchall()
    conn.close()
    return data

# Queries for shows and movies
shows_query = """
SELECT 
    Series.Path,
    SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3), '\\') - 1) as vgName,
    SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + LENGTH(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3), '\\') - 1)) + 4) as showFolderName,
    'son.' || Series.Id as Id, 
    Series.Title, 
    Series.Year, 
    Series.FirstAired, 
    Series.LastAired, 
    Series.lastInfoSync, 
    ROUND(SUM(EpisodeFiles.Size) / (1024 * 1024 * 1024.0), 1) as TotalSizeGB,
    SUBSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + LENGTH(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3), '\\') - 1)) + 4), INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + LENGTH(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3), '\\') - 1)) + 4), '(') + 1, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + LENGTH(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3), '\\') - 1)) + 4), ')') - INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + LENGTH(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3, INSTR(SUBSTR(Series.Path, INSTR(Series.Path, 'VG\\') + 3), '\\') - 1)) + 4), '(') - 1) as showFolderYYYY
FROM Series
LEFT JOIN EpisodeFiles ON Series.Id = EpisodeFiles.SeriesId
GROUP BY 
    Series.Path,
    vgName,
    showFolderName,
    Series.Id, 
    Series.Title, 
    Series.Year, 
    Series.FirstAired, 
    Series.LastAired, 
    Series.lastInfoSync
"""

movies_query = """
SELECT 
    movies.path,
    SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3), '\\') - 1) as vgName,
    SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + LENGTH(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3), '\\') - 1)) + 4) as movieFolderName,
    'rad.' || movies.id as Id, 
    movieMetaData.title, 
    movieMetaData.year, 
    movieMetaData.PhysicalRelease as FirstAired, 
    movieMetaData.DigitalRelease as LastAired, 
    movieMetaData.lastInfoSync, 
    ROUND(SUM(movieFiles.size) / (1024 * 1024 * 1024.0), 1) as TotalSizeGB,
    SUBSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + LENGTH(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3), '\\') - 1)) + 4), INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + LENGTH(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3), '\\') - 1)) + 4), '(') + 1, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + LENGTH(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3), '\\') - 1)) + 4), ')') - INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + LENGTH(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3, INSTR(SUBSTR(movies.Path, INSTR(movies.Path, 'VG\\') + 3), '\\') - 1)) + 4), '(') - 1) as movieFolderYYYY
FROM movies
LEFT JOIN movieMetaData ON movies.MovieMetadataId = movieMetaData.Id
LEFT JOIN movieFiles ON movies.MovieFileId = movieFiles.Id
GROUP BY 
    movies.path,
    vgName,
    movieFolderName,
    movies.id, 
    movieMetaData.title, 
    movieMetaData.year
"""

# Fetch data from the databases
shows_data = fetch_data(shows_query, r'C:\Users\All Users\Sonarr\sonarr.db')
movies_data = fetch_data(movies_query, r'C:\Users\All Users\Radarr\radarr.db')

# Combine the data
combined_data = shows_data + movies_data

# Create a new Excel workbook and add a worksheet
logging.info('Creating Excel workbook...')
wb = Workbook()
ws = wb.active
ws.title = "media_dumpARR"

# Add headers to the worksheet
headers = ["Path", "FolderName", "Id", "Title", "Year", "FirstAired", "LastAired", "lastInfoSync", "TotalSizeGB", "vgName", "FolderYYYY", "Backup Check"]
ws.append(headers)

# Add combined data to the worksheet
logging.info('Adding combined data to the worksheet...')
for i, row in enumerate(combined_data, start=2):  # start=2 to account for header row
    path, vgName, folderName, id, title, year, firstAired, lastAired, lastInfoSync, totalSizeGB, folderYYYY = row
    reordered_row = [path, folderName, id, title, year, firstAired, lastAired, lastInfoSync, totalSizeGB, vgName, folderYYYY]
    ws.append(reordered_row)
    formula = f"=IFERROR(VLOOKUP(A{i},'[backup_control.xlsx]Media-Folders'!$A:$E,5,FALSE),\"-\")"
    ws[f'L{i}'] = formula  # Assuming 'L' is the last column

# Create a summary dataset grouped by vgName
logging.info('Creating summary dataset grouped by vgName...')
summary_data = {}
for row in combined_data:
    vg_name = row[1]  # Assuming vgName is the 2nd column in the result set
    total_size_gb = float(row[9]) if row[9] is not None else 0.0  # Handle None values and convert to float
    first_aired = row[6] if row[6] is not None else '1900-01-01'  # Handle None values
    last_aired = row[7] if row[7] is not None else '1900-01-01'  # Handle None values
    year = row[5] if row[5] is not None else 1900  # Handle None values
    if vg_name not in summary_data:
        summary_data[vg_name] = {
            'count': 0,
            'sum': 0.0,
            'min_first_aired': first_aired,
            'max_first_aired': first_aired,
            'min_last_aired': last_aired,
            'max_last_aired': last_aired,
            'min_year': year,
            'max_year': year
        }
    summary_data[vg_name]['count'] += 1
    summary_data[vg_name]['sum'] += total_size_gb
    summary_data[vg_name]['min_first_aired'] = min(summary_data[vg_name]['min_first_aired'], first_aired)
    summary_data[vg_name]['max_first_aired'] = max(summary_data[vg_name]['max_first_aired'], first_aired)
    summary_data[vg_name]['min_last_aired'] = min(summary_data[vg_name]['min_last_aired'], last_aired)
    summary_data[vg_name]['max_last_aired'] = max(summary_data[vg_name]['max_last_aired'], last_aired)
    summary_data[vg_name]['min_year'] = min(summary_data[vg_name]['min_year'], year)
    summary_data[vg_name]['max_year'] = max(summary_data[vg_name]['max_year'], year)

# Create a new sheet for the summary data
summary_sheet_name = 'summary-data'
logging.info(f'Creating new sheet: {summary_sheet_name}')
summary_sheet = wb.create_sheet(title=summary_sheet_name)

# Write summary column names to the sheet
summary_column_names = ['vgName', 'count', 'sum', 'min_first_aired', 'max_first_aired', 'min_last_aired', 'max_last_aired', 'min_year', 'max_year']
logging.info('Writing summary column names to the Excel sheet...')
summary_sheet.append(summary_column_names)

# Write summary data to the sheet
logging.info('Writing summary data to the Excel sheet...')
for vg_name, data in summary_data.items():
    summary_row = [
        vg_name,
        data['count'],
        data['sum'],
        data['min_first_aired'],
        data['max_first_aired'],
        data['min_last_aired'],
        data['max_last_aired'],
        data['min_year'],
        data['max_year']
    ]
    summary_sheet.append(summary_row)

# Save the workbook to a file
excel_file = r'C:\media.tt.omp\metadata\media_dumpARR.xlsx'
logging.info(f'Saving Excel file: {excel_file}')
try:
    wb.save(excel_file)
    logging.info('Workbook saved successfully.')
except PermissionError:
    logging.error(f"\033[91mPermission denied: Unable to save the file '{excel_file}'. Please close the file if it's open and try again.\033[0m")
logging.info('Script completed successfully.')